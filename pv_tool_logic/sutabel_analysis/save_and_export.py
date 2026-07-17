"""
Save and Export functies voor Sutabel-m analyse.

Deze module bevat functies voor het opslaan en exporteren van sutabel-m analyseresultaten
naar Excel en PDF-formaat.
"""

from typing import TYPE_CHECKING, List
from pandas import ExcelWriter, concat, DataFrame, read_excel
from datetime import datetime
import xlwings as xw
import importlib.resources
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, LongTable, PageBreak
from pv_tool_logic.imports.excel_utils import format_excel_sheet
import plotly.graph_objects as go
from pathlib import Path

try:
    from reportlab.platypus import Image as RLImage
    from PIL import Image as PILImage

    PIL_AVAILABLE = True
except ImportError:
    RLImage = None
    PIL_AVAILABLE = False
    PILImage = None
    print("Waarschuwing: PIL (Pillow) is niet beschikbaar. Figuren kunnen niet aan PDF worden toegevoegd.")

if TYPE_CHECKING:
    from pv_tool_logic.sutabel_analysis.sutabel_analysis import SUTABEL


def add_sutabel_results_to_dbase(self: "SUTABEL", path: str, file_name: str = "Template_PVtool5_0.xlsx"):
    # TODO wordt vervangen door add_results_to_template
    """
    Voegt de sutabel-m analyseresultaten toe aan de database Excel-bestand.

    Parameters
    ----------
    self: sutabel
        Instantie van de sutabel analyse klasse
    path: str
        Pad naar de map waar het Excel-bestand staat
    file_name : str
        Naam van het Excel-bestand

    Returns
    -------
    DataFrame
        Bijgewerkte DataFrame met alle resultaten
    """
    file_path = f"{path}/{file_name}"

    if self.sutabel_grafiek is None or self.e_a1_sutabel is None:
        self._run_sutabel()
        if hasattr(self, "vc_fit_kar_sutabel") and self.vc_fit_kar_sutabel is not None:
            self.get_sutabel_parameters(vc_fit_kar_sutabel=self.vc_fit_kar_sutabel)
        else:
            self.get_sutabel_parameters()

    expected_columns = [
        "PV_NAAM",
        "PV_REK",
        "PV_TYPE_PROEF",
        "PV_ANALYSE",
        "PV_RESULTAAT_ID",
        "PV_TYPEVERZAMELING",
        "PV_e_a1_GEM",
        "PV_e_a2_GEM",
        "PV_svgm_GEM",
        "PV_m_GEM",
        "PV_a1_KAR",
        "PV_a2_KAR",
        "PV_svgm_KAR",
        "PV_m_KAR",
        "PV_vc_FIT_KAR",
        "PV_STDEV_LOGN_vc",
        "PV_STEYX",
        "PV_VGWNAT_GEM",
        "PV_VGWNAT_SD",
        "PV_WATERGEHALTE_GEM",
        "PV_WATERGEHALTE_SD",
        "Timestamp",
    ]

    new_row = {
        "PV_NAAM": self.investigation_groups[0],
        "PV_REK": self.effective_stress,
        "PV_TYPE_PROEF": self.analysis_type.split("_")[0],
        "PV_ANALYSE": "_".join(self.analysis_type.split("_")[1:]),
        "PV_RESULTAAT_ID": f"{self.investigation_groups[0]}_{self.effective_stress}_{self.analysis_type}",
        "PV_TYPEVERZAMELING": self.alpha,
        "PV_e_a1_GEM": round(self.e_a1_sutabel, 6) if self.e_a1_sutabel is not None else None,
        "PV_e_a2_GEM": round(self.e_a2_sutabel, 6) if self.e_a2_sutabel is not None else None,
        "PV_svgm_GEM": round(self.svgm_gem_sutabel, 6) if self.svgm_gem_sutabel is not None else None,
        "PV_m_GEM": round(self.m_gem_sutabel, 6) if self.m_gem_sutabel is not None else None,
        "PV_a1_KAR": round(self.a1_kar_sutabel, 6) if self.a1_kar_sutabel is not None else None,
        "PV_a2_KAR": round(self.a2_kar_sutabel, 6) if self.a2_kar_sutabel is not None else None,
        "PV_svgm_KAR": round(self.svgm_kar_sutabel, 6) if self.svgm_kar_sutabel is not None else None,
        "PV_m_KAR": round(self.m_kar_sutabel, 6) if self.m_kar_sutabel is not None else None,
        "PV_vc_FIT_KAR": round(self.vc_fit_kar_sutabel, 6) if self.vc_fit_kar_sutabel is not None else None,
        "PV_STDEV_LOGN_vc": (round(self.STDEV_logn_vc_sutabel, 6) if self.STDEV_logn_vc_sutabel is not None else None),
        "PV_STEYX": round(self.steyx_sutabel, 6) if self.steyx_sutabel is not None else None,
        "PV_VGWNAT_GEM": round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
        "PV_VGWNAT_SD": round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
        "PV_WATERGEHALTE_GEM": round(self.calc_watergehalte_gem, 3) if self.calc_watergehalte_gem is not None else None,
        "PV_WATERGEHALTE_SD": round(self.calc_watergehalte_sd, 3) if self.calc_watergehalte_sd is not None else None,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    workbook = load_workbook(file_path)

    if "Resultaten SU-tabel-m" in workbook.sheetnames:
        df_existing = read_excel(file_path, sheet_name="Resultaten SU-tabel-m")
        df_existing = df_existing.dropna(how="all")
        df_existing.columns = df_existing.columns.astype(str)
        new_row_df = DataFrame([new_row], columns=df_existing.columns)
        df_updated = concat([df_existing, new_row_df], ignore_index=True)
    else:
        df_updated = DataFrame([new_row], columns=expected_columns)

    # Ensure all column headers are strings
    df_updated.columns = df_updated.columns.astype(str)

    # Write data to Excel
    with ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        df_updated.to_excel(writer, sheet_name="Resultaten SU-tabel-m", index=False)

    # Format the Excel sheet
    format_excel_sheet(
        file_path=file_path,
        sheet_name="Resultaten SU-tabel-m",
        num_columns=df_updated.shape[1],
        num_rows=df_updated.shape[0],
        table_name="ResultatenSUTabelMTable",
        index=False,
    )

    print(f"Sutabel resultaten toegevoegd aan database: {file_path}")
    return df_updated


def add_results_to_template(self: "SUTABEL", path, export_name=None):
    """
    Voegt de sutabel-m analyseresultaten toe aan de database Excel-bestand.

    Parameters
    ----------
    self: sutabel
        Instantie van de sutabel analyse klasse
    path: str
        Pad naar de map waar het Excel-bestand staat
    export_name
        Naam van het Excel-bestand

    Returns
    -------
    DataFrame
        Bijgewerkte DataFrame met alle resultaten
    """
    if export_name is None:
        export_name = "Template_PVtool5_0.xlsx"
    file_path = (Path(path) / export_name).resolve()
    sheet_name = "Resultaten SU-tabel-m"

    if self.sutabel_grafiek is None or self.e_a1_sutabel is None:
        self._run_sutabel()
        if hasattr(self, "vc_fit_kar_sutabel") and self.vc_fit_kar_sutabel is not None:
            self.get_sutabel_parameters(vc_fit_kar_sutabel=self.vc_fit_kar_sutabel)
        else:
            self.get_sutabel_parameters()

    expected_columns = [
        "PV_NAAM",
        "PV_REK",
        "PV_TYPE_PROEF",
        "PV_ANALYSE",
        "PV_RESULTAAT_ID",
        "PV_TYPEVERZAMELING",
        "PV_e_a1_GEM",
        "PV_e_a2_GEM",
        "PV_svgm_GEM",
        "PV_m_GEM",
        "PV_a1_KAR",
        "PV_a2_KAR",
        "PV_svgm_KAR",
        "PV_m_KAR",
        "PV_vc_FIT_KAR",
        "PV_STDEV_LOGN_vc",
        "PV_STEYX",
        "PV_VGWNAT_GEM",
        "PV_VGWNAT_SD",
        "PV_WATERGEHALTE_GEM",
        "PV_WATERGEHALTE_SD",
        "Timestamp",
        "SU-tabel s'v 1",
        "SU-tabel s'v 2",
        "SU-tabel s'v 3",
        "SU-tabel s'v 4",
        "SU-tabel s'v 5",
        "SU-tabel s'v 6",
        "SU-tabel s'v 7",
        "SU-tabel su_gem 1",
        "SU-tabel su_gem 2",
        "SU-tabel su_gem 3",
        "SU-tabel su_gem 4",
        "SU-tabel su_gem 5",
        "SU-tabel su_gem 6",
        "SU-tabel su_gem 7",
        "SU-tabel su_kar 1",
        "SU-tabel su_kar 2",
        "SU-tabel su_kar 3",
        "SU-tabel su_kar 4",
        "SU-tabel su_kar 5",
        "SU-tabel su_kar 6",
        "SU-tabel su_kar 7",
    ]

    new_row = {
        "PV_NAAM": self.investigation_groups[0],
        "PV_REK": self.effective_stress,
        "PV_TYPE_PROEF": self.analysis_type.split("_")[0],
        "PV_ANALYSE": "_".join(self.analysis_type.split("_")[1:]),
        "PV_RESULTAAT_ID": f"{self.investigation_groups[0]}_{self.effective_stress}_{self.analysis_type}",
        "PV_TYPEVERZAMELING": self.alpha,
        "PV_e_a1_GEM": round(self.e_a1_sutabel, 3) if self.e_a1_sutabel is not None else None,
        "PV_e_a2_GEM": round(self.e_a2_sutabel, 3) if self.e_a2_sutabel is not None else None,
        "PV_svgm_GEM": round(self.svgm_gem_sutabel, 3) if self.svgm_gem_sutabel is not None else None,
        "PV_m_GEM": round(self.m_gem_sutabel, 3) if self.m_gem_sutabel is not None else None,
        "PV_a1_KAR": round(self.a1_kar_sutabel, 3) if self.a1_kar_sutabel is not None else None,
        "PV_a2_KAR": round(self.a2_kar_sutabel, 3) if self.a2_kar_sutabel is not None else None,
        "PV_svgm_KAR": round(self.svgm_kar_sutabel, 3) if self.svgm_kar_sutabel is not None else None,
        "PV_m_KAR": round(self.m_kar_sutabel, 3) if self.m_kar_sutabel is not None else None,
        "PV_vc_FIT_KAR": round(self.vc_fit_kar_sutabel, 3) if self.vc_fit_kar_sutabel is not None else None,
        "PV_STDEV_LOGN_vc": (round(self.STDEV_logn_vc_sutabel, 3) if self.STDEV_logn_vc_sutabel is not None else None),
        "PV_STEYX": round(self.steyx_sutabel, 3) if self.steyx_sutabel is not None else None,
        "PV_VGWNAT_GEM": round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
        "PV_VGWNAT_SD": round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
        "PV_WATERGEHALTE_GEM": round(self.calc_watergehalte_gem, 3) if self.calc_watergehalte_gem is not None else None,
        "PV_WATERGEHALTE_SD": round(self.calc_watergehalte_sd, 3) if self.calc_watergehalte_sd is not None else None,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "SU-tabel s'v 1": (round(self.sutabel_grafiek["s'v"].iloc[0], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 2": (round(self.sutabel_grafiek["s'v"].iloc[1], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 3": (round(self.sutabel_grafiek["s'v"].iloc[2], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 4": (round(self.sutabel_grafiek["s'v"].iloc[3], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 5": (round(self.sutabel_grafiek["s'v"].iloc[4], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 6": (round(self.sutabel_grafiek["s'v"].iloc[5], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel s'v 7": (round(self.sutabel_grafiek["s'v"].iloc[6], 3) if self.sutabel_grafiek is not None else None),
        "SU-tabel su_gem 1": (
            round(self.sutabel_grafiek["su_gem"].iloc[0], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 2": (
            round(self.sutabel_grafiek["su_gem"].iloc[1], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 3": (
            round(self.sutabel_grafiek["su_gem"].iloc[2], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 4": (
            round(self.sutabel_grafiek["su_gem"].iloc[3], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 5": (
            round(self.sutabel_grafiek["su_gem"].iloc[4], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 6": (
            round(self.sutabel_grafiek["su_gem"].iloc[5], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_gem 7": (
            round(self.sutabel_grafiek["su_gem"].iloc[6], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 1": (
            round(self.sutabel_grafiek["su_kar"].iloc[0], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 2": (
            round(self.sutabel_grafiek["su_kar"].iloc[1], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 3": (
            round(self.sutabel_grafiek["su_kar"].iloc[2], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 4": (
            round(self.sutabel_grafiek["su_kar"].iloc[3], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 5": (
            round(self.sutabel_grafiek["su_kar"].iloc[4], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 6": (
            round(self.sutabel_grafiek["su_kar"].iloc[5], 3) if self.sutabel_grafiek is not None else None
        ),
        "SU-tabel su_kar 7": (
            round(self.sutabel_grafiek["su_kar"].iloc[6], 3) if self.sutabel_grafiek is not None else None
        ),
    }

    header_row = 7

    app = None
    wb = None
    opened_by_function = False

    try:
        # Zoek eerst naar reeds geopend workbook
        for app_instance in xw.apps:
            for wb_instance in app_instance.books:
                try:
                    if Path(wb_instance.fullname).resolve() == file_path:
                        wb = wb_instance
                        break
                except Exception:
                    pass

            if wb is not None:
                break

        # Niet open -> onzichtbaar openen
        if wb is None:
            app = xw.App(visible=False)
            app.display_alerts = False
            app.screen_updating = False

            if file_path.exists():
                wb = app.books.open(str(file_path))
            else:
                with importlib.resources.path(
                        "pv_tool_logic.templates",
                        "Template_PVtool5_0.xlsx"
                ) as template_path:

                    wb = app.books.open(str(template_path))
                    wb.save(str(file_path))

            opened_by_function = True

        # Sheet ophalen of aanmaken
        sheet_names = [sheet.name for sheet in wb.sheets]

        if sheet_name in sheet_names:
            ws = wb.sheets[sheet_name]
        else:
            ws = wb.sheets.add(sheet_name, after=wb.sheets[-1])

            # headers wegschrijven op rij 7
            ws.range((header_row, 1)).value = [expected_columns]

        # Headers controleren
        if ws.range((header_row, 1)).value in [None, ""]:
            ws.range((header_row, 1)).value = [expected_columns]

        # Bestaande kolommen uitlezen
        last_col = ws.range((header_row, ws.cells.last_cell.column)).end("left").column

        col_name_to_idx = {
            ws.range((header_row, col)).value: col
            for col in range(1, last_col + 1)
            if ws.range((header_row, col)).value is not None
        }

        # Eerste lege rij bepalen
        last_row = ws.range(
            "A" + str(ws.cells.last_cell.row)
        ).end("up").row

        first_empty_row = max(last_row + 1, header_row + 1)

        # Resultaatregel schrijven
        for col_name, col_idx in col_name_to_idx.items():
            ws.range((first_empty_row, col_idx)).value = new_row.get(
                col_name,
                ""
            )

        wb.save()

        print(
            f"Resultaat toegevoegd aan template in tabblad '{sheet_name}'."
        )

    finally:

        if opened_by_function and wb is not None:
            wb.close()

        if opened_by_function and app is not None:
            app.quit()


def _create_sutabel_input_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de invoerselectie informatie voor sutabel analyse.

    Returns
    -------
    Table
        ReportLab tabel object met de invoerselectie informatie
    """
    if self.sutabel_data_df is None:
        return Table([["Geen invoerdata beschikbaar"]], hAlign="LEFT")

    columns_base = [
        "PV_NAAM",
        "BORING_POSITIE",
        "MONSTER_NIVEAU_NAP_VANAF",
        "MONSTER_NIVEAU_NAP_TOT",
        "ANA_TERREINSPANNING",
        "ANA_GRENSSPANNING_REKEN",
        "ANA_POP_VELD",
    ]

    if self.analysis_type.startswith("TXT"):
        columns_extra = ["TXT_SS_VOLUMEGEWICHT_NAT", "TXT_SS_WATERGEHALTE_VOOR"]
    else:
        columns_extra = ["DSS_VOLUMEGEWICHT_NAT", "DSS_WATERGEHALTE_VOOR"]

    available_columns = [col for col in columns_base + columns_extra if col in self.total_sutabel_data_df.columns]

    if not available_columns:
        return Table([["Geen invoerdata kolommen beschikbaar"]], hAlign="LEFT")

    table_df = self.total_sutabel_data_df[available_columns].copy()

    additional_columns = ["S'v", "Su", "consolidatietype"]
    for col in additional_columns:
        if col in self.sutabel_data_df.columns and col not in table_df.columns:
            # Zorg ervoor dat de indices overeenkomen
            if len(table_df) == len(self.sutabel_data_df):
                table_df[col] = self.sutabel_data_df[col].values

    column_mapping = {
        "PV_NAAM": "Groep",
        "BORING_POSITIE": "Pos.",
        "MONSTER_NIVEAU_NAP_VANAF": "NAP Van\n[m]",
        "MONSTER_NIVEAU_NAP_TOT": "NAP Tot\n[m]",
        "TXT_SS_VOLUMEGEWICHT_NAT": "VGW nat\n[kN/m3]",
        "TXT_SS_VOLUMEGEWICHT_DRG": "VGW drg\n[kN/m3]",
        "TXT_SS_WATERGEHALTE_VOOR": "Water\n[%]",
        "DSS_VOLUMEGEWICHT_NAT": "VGW nat\n[kN/m3]",
        "DSS_VOLUMEGEWICHT_DRG": "VGW drg\n[kN/m3]",
        "DSS_WATERGEHALTE_VOOR": "Water\n[%]",
        "S'v": "σ'v\n[kPa]",
        "Su": "su\n[kPa]",
        "OCR": "OCR\n[-]",
        "ANA_TERREINSPANNING": "Terreinspanning\n[kPa]",
        "ANA_GRENSSPANNING_REKEN": "Grensspanning\n[kPa]",
        "ANA_POP_VELD": "POP\n[-]",
    }

    table_df = table_df.rename(columns={k: v for k, v in column_mapping.items() if k in table_df.columns})

    # Kolom-specifieke afronding; niet-numerieke waarden worden ongemoeid gelaten
    vgw_col = "VGW nat\n[kN/m3]"
    water_col = "Water\n[%]"

    def round_mixed(series, decimals: int):
        """Rondt numerieke waarden af; niet-numerieke waarden blijven ongewijzigd."""

        def _round(val):
            try:
                return round(float(val), decimals)
            except (TypeError, ValueError):
                return val

        return series.apply(_round)

    for col in table_df.columns:
        if col == vgw_col:
            table_df[col] = round_mixed(table_df[col], 2)
        elif col == water_col:
            table_df[col] = round_mixed(table_df[col], 1)
        elif table_df[col].dtype in ["float64", "float32"]:
            table_df[col] = table_df[col].round(3)

    from reportlab.platypus import Paragraph

    header_style = getSampleStyleSheet()["Normal"]
    header_style.fontSize = 7
    header_style.fontName = "Helvetica-Bold"
    header_style.alignment = 0  # Left align

    # Create header row with index column
    header_paragraphs = [Paragraph("Monster ID", header_style)]

    for col in table_df.columns:
        col_display = col.replace("_", "_<br/>")
        header_paragraphs.append(Paragraph(col_display, header_style))

    # Create data rows with index
    data_rows = []
    for idx, row in table_df.iterrows():
        data_row = [str(idx)] + [str(val) if val != "" else "" for val in row]
        data_rows.append(data_row)

    t_data = [header_paragraphs] + data_rows

    t1 = LongTable(t_data, repeatRows=1, hAlign="LEFT")
    t1.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("ROWBACKGROUNDS", (0, 0), (-1, 0), [colors.lightgrey]),
            ]
        )
    )
    return t1


def _create_sutabel_parameters_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de sutabel-m parameters.

    Returns
    -------
    Table
        ReportLab tabel object met de sutabel-m parameters
    """
    parameters: List[list] = []

    if hasattr(self, "e_a1_sutabel") and self.e_a1_sutabel is not None:
        parameters.append(["e_a1 (snijpunt gemiddeld) [-]", f"{self.e_a1_sutabel:.3f}"])
    if hasattr(self, "e_a2_sutabel") and self.e_a2_sutabel is not None:
        parameters.append(["e_a2 (helling gemiddeld) [-]", f"{self.e_a2_sutabel:.3f}"])
    if hasattr(self, "svgm_gem_sutabel") and self.svgm_gem_sutabel is not None:
        parameters.append(["svgm_gem [kPa]", f"{self.svgm_gem_sutabel:.3f}"])
    if hasattr(self, "m_gem_sutabel") and self.m_gem_sutabel is not None:
        parameters.append(["m_gem [-]", f"{self.m_gem_sutabel:.3f}"])

    if hasattr(self, "a1_kar_sutabel") and self.a1_kar_sutabel is not None:
        parameters.append(["a1_kar (snijpunt karakteristiek) [-]", f"{self.a1_kar_sutabel:.3f}"])
    if hasattr(self, "a2_kar_sutabel") and self.a2_kar_sutabel is not None:
        parameters.append(["a2_kar (helling karakteristiek) [-]", f"{self.a2_kar_sutabel:.3f}"])
    if hasattr(self, "svgm_kar_sutabel") and self.svgm_kar_sutabel is not None:
        parameters.append(["svgm_kar [kPa]", f"{self.svgm_kar_sutabel:.3f}"])
    if hasattr(self, "m_kar_sutabel") and self.m_kar_sutabel is not None:
        parameters.append(["m_kar [-]", f"{self.m_kar_sutabel:.3f}"])

    if hasattr(self, "vc_fit_kar_sutabel") and self.vc_fit_kar_sutabel is not None:
        parameters.append(["vc_fit_kar [-]", f"{self.vc_fit_kar_sutabel:.3f}"])
    if hasattr(self, "STDEV_logn_vc_sutabel") and self.STDEV_logn_vc_sutabel is not None:
        parameters.append(["STDEV lognormaal [-]", f"{self.STDEV_logn_vc_sutabel:.3f}"])
    if hasattr(self, "steyx_sutabel") and self.steyx_sutabel is not None:
        parameters.append(["STEYX [-]", f"{self.steyx_sutabel:.3f}"])

    parameters.append(["Type verzameling: lokaal = 1.0; regionaal = 0.75", f"{self.alpha:.2f}"])

    if hasattr(self, "calc_vgwnat_gem") and self.calc_vgwnat_gem is not None:
        parameters.append(["VGW nat gemiddeld [kN/m3]", f"{self.calc_vgwnat_gem:.2f}"])
    if hasattr(self, "calc_watergehalte_gem") and self.calc_watergehalte_gem is not None:
        parameters.append(["Watergehalte gemiddeld [%]", f"{self.calc_watergehalte_gem:.2f}"])

    t = Table([["Parameter", "Waarde"]] + parameters, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )
    return t


def _create_sutabel_grafiek_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de sutabel grafiek data (su_gem en su_kar lijnen).

    Returns
    -------
    Table
        ReportLab tabel object met de sutabel grafiek data
    """
    if self.sutabel_grafiek is None:
        return Table([["Geen sutabel grafiek data beschikbaar"]], hAlign="LEFT")

    df = self.sutabel_grafiek.copy()
    for col in df.columns:
        if df[col].dtype in ["float64", "float32"]:
            df[col] = df[col].round(3)

    header = df.columns.tolist()
    data = df.values.tolist()
    t_data = [header] + data

    t = Table(t_data, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )
    return t


def _create_su_fit_vc_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de su fit constante vc data.

    Returns
    -------
    Table
        ReportLab tabel object met de su fit vc data
    """
    if self.su_fit_constante_vc is None:
        return Table([["Geen vc fit data beschikbaar (vc_fit_kar niet opgegeven)"]], hAlign="LEFT")

    df = self.su_fit_constante_vc.copy()
    for col in df.columns:
        if df[col].dtype in ["float64", "float32"]:
            df[col] = df[col].round(3)

    header = df.columns.tolist()
    data = df.values.tolist()
    t_data = [header] + data

    t = Table(t_data, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )
    return t


def save_sutabel_to_pdf(self: "SUTABEL", path: str, vc_fit_kar: float = None) -> str:
    """
    Slaat de sutabel-m analyseresultaten op in een PDF-document.

    De PDF bevat:
    - Titel met analysedetails
    - Beide overzichtsfiguren (ln(s'v)-ln(su) en s'v-su)
    - Tabel met parameters
    - Tabel met sutabel grafiek data (su_gem en su_kar)
    - Tabel met su fit constante vc data (indien van toepassing)
    - Tabel met invoerselectie informatie

    Parameters
    ----------
    self: SUTABEL
        Instantie van de SUTABEL klasse met de analyse uitvoert
    path: str
        Map locatie waar het PDF-bestand moet worden opgeslagen
    vc_fit_kar: float, optioneel
        Coefficient of Variation voor de fit

    Returns
    -------
    str
        Het absolute bestandspad van het aangemaakte PDF-bestand
    """
    title = f"Sutabel-m analyse met {self.effective_stress} op {self.investigation_groups[0]}"
    file_name = (
        f"sutabel_pdf_export_{self.investigation_groups[0]}_{self.analysis_type}_"
        f"{str(self.effective_stress).replace('%', 'procent_').replace(' ', '')}.pdf"
    )
    file_path = f"{path}/{file_name}"

    self._run_sutabel()
    if vc_fit_kar is not None:
        self.get_sutabel_parameters(vc_fit_kar_sutabel=vc_fit_kar)
    elif hasattr(self, "vc_fit_kar_sutabel") and self.vc_fit_kar_sutabel is not None:
        self.get_sutabel_parameters(vc_fit_kar_sutabel=self.vc_fit_kar_sutabel)
    else:
        self.get_sutabel_parameters()

    if self.sutabel_grafiek is None:
        self.calculate_sutabel_grafiek()

    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    if "Left" not in styles:
        styles.add(ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT))
    if "TitleLeft" not in styles:
        styles.add(ParagraphStyle(name="TitleLeft", parent=styles["Title"], alignment=TA_LEFT))
    if "Heading3" not in styles:
        styles.add(ParagraphStyle(name="Heading3", parent=styles["Heading2"], alignment=TA_LEFT))

    story = [Paragraph(title, styles["TitleLeft"]), Spacer(width=1, height=12)]

    if not PIL_AVAILABLE:
        story.append(
            Paragraph("Figuren kunnen niet worden toegevoegd: PIL (Pillow) niet beschikbaar", styles["Normal"])
        )
        story.append(Spacer(width=1, height=12))
    else:
        try:
            fig_path1 = f"{path}/temp_sutabel_plot1.png"
            self.show_title = False

            self.figure_sv_su = go.Figure()
            self.set_figure_sv_su_sutabel()

            if self.figure_sv_su is not None and len(self.figure_sv_su.data) > 0:
                fig_width = 1280
                fig_height = 720
                self.figure_sv_su.write_image(fig_path1, width=fig_width, height=fig_height, scale=4, format="png")

                title_and_heading_height = 100
                available_height = doc.height - title_and_heading_height

                with PILImage.open(fig_path1) as im:
                    img_width_px, img_height_px = im.size

                max_width_pt = doc.width * 0.95
                aspect = img_height_px / img_width_px

                img_width_pt = min(max_width_pt, doc.width)
                img_height_pt = img_width_pt * aspect

                if img_height_pt > available_height:
                    img_height_pt = available_height * 0.95
                    img_width_pt = img_height_pt / aspect

                # Maak ReportLab Image aan
                img1 = RLImage(fig_path1)
                img1.drawWidth = img_width_pt
                img1.drawHeight = img_height_pt
                img1.hAlign = "LEFT"

                story.append(Paragraph("SU-tabel (σ'<sub>v</sub> - s<sub>u</sub> relatie)", styles["Heading2"]))
                story.append(Spacer(width=1, height=6))
                story.append(img1)
                story.append(PageBreak())
            else:
                story.append(Paragraph("Figuur 1 kon niet worden gegenereerd", styles["Normal"]))
                story.append(Spacer(width=1, height=12))
        except Exception as e:
            story.append(Paragraph(f"Fout bij genereren Figuur 1: {str(e)}", styles["Normal"]))
            story.append(Spacer(width=1, height=12))

        try:
            fig_path2 = f"{path}/temp_sutabel_plot2.png"
            self.show_title = False

            self.figure_ln_sv_ln_su = go.Figure()
            self.set_figure_ln_sv_ln_su_sutabel()

            if self.figure_ln_sv_ln_su is not None and len(self.figure_ln_sv_ln_su.data) > 0:
                fig_width = 1280
                fig_height = 720
                self.figure_ln_sv_ln_su.write_image(
                    fig_path2, width=fig_width, height=fig_height, scale=4, format="png"
                )

                with PILImage.open(fig_path2) as im:
                    img_width_px, img_height_px = im.size

                max_width_pt = doc.width * 0.95
                aspect = img_height_px / img_width_px

                img_width_pt = min(max_width_pt, doc.width)
                img_height_pt = img_width_pt * aspect

                available_height = doc.height - 50

                if img_height_pt > available_height:
                    img_height_pt = available_height * 0.95
                    img_width_pt = img_height_pt / aspect

                img2 = RLImage(fig_path2)
                img2.drawWidth = img_width_pt
                img2.drawHeight = img_height_pt
                img2.hAlign = "LEFT"

                story.append(Paragraph("SU - tabel logaritmisch (LN(s'vc) - LN(su))", styles["Heading2"]))
                story.append(Spacer(width=1, height=6))
                story.append(img2)
                story.append(PageBreak())
            else:
                story.append(Paragraph("Figuur 2 kon niet worden gegenereerd", styles["Normal"]))
                story.append(Spacer(width=1, height=12))
        except Exception as e:
            story.append(Paragraph(f"Fout bij genereren Figuur 2: {str(e)}", styles["Normal"]))
            story.append(Spacer(width=1, height=12))

    # Voeg parameters tabel toe
    story.append(Paragraph("Parameters", styles["Heading2"]))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_parameters_table(self))
    story.append(Spacer(width=1, height=12))

    # Voeg sutabel grafiek data toe
    story.append(Paragraph("Sutabel grafiek data (su_gem en su_kar lijnen)", styles["Heading2"]))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_grafiek_table(self))
    story.append(Spacer(width=1, height=12))

    # Voeg su fit vc data toe (indien beschikbaar)
    if self.su_fit_constante_vc is not None:
        story.append(Paragraph("Su fit met constante vc data", styles["Heading2"]))
        story.append(Spacer(width=1, height=6))
        story.append(_create_su_fit_vc_table(self))
        story.append(Spacer(width=1, height=12))

    story.append(PageBreak())
    story.append(Paragraph("Invoerselectie", styles["Heading2"]))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_input_table(self))

    try:
        doc.build(story)
        print(f"Sutabel PDF export voltooid: {file_path}")

        # Ruim tijdelijke plot bestanden op
        import os

        temp_files = [f"{path}/temp_sutabel_plot1.png", f"{path}/temp_sutabel_plot2.png"]
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                print(f"Waarschuwing: Kon tijdelijk plot bestand niet verwijderen {temp_file}: {e}")

        return file_path
    except Exception as e:
        print(f"Fout bij maken PDF: {e}")
        raise
