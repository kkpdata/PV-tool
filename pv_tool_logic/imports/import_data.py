import os
import platform
from openpyxl import load_workbook

from pandas import DataFrame
import pandas as pd
import importlib.resources
from typing import Optional, Literal
from pathlib import Path
import importlib.resources
import xlwings as xw

from pv_tool_logic.imports.add_ana_columns import recalc_alg_boring_monsternr
from pv_tool_logic.imports.create_dbase import (
    add_missing_columns,
    select_columns,
    alg_columns,
    add_ana_columns,
    add_pv_naam,
)
from pv_tool_logic.imports.import_options import import_dbase, import_pv_tool, import_stowa
from pv_tool_logic.imports.validation import Validation
from pv_tool_logic.imports.globals import PV_TOOL_DBASE_COLUMNS, ANA_COLUMNS

class Dbase:
    """Deze class bevat alle functies die te maken hebben met het bouwen de Dbase-dataframe"""

    def __init__(self):
        self.stowa_df: Optional[DataFrame] = None
        self.pv_tool: Optional[DataFrame] = None
        self.dbase_df: Optional[DataFrame] = None
        self.validation = Validation(dbase=self)

    def _create_dbase(self, source: Literal["Stowa", "PV-tool", "Dbase"]):
        """Maakt de dbase-dataframe"""
        if source == "Stowa":
            add_missing_columns(self)
            alg_columns(self)
            add_ana_columns(self)
            recalc_alg_boring_monsternr(self)
            add_pv_naam(self)
        elif source == "PV-tool":
            select_columns(self)
            alg_columns(self)
            add_ana_columns(self)
            recalc_alg_boring_monsternr(self)
            add_pv_naam(self)
        elif source == "Dbase":
            add_ana_columns(self)
            recalc_alg_boring_monsternr(self)
            add_pv_naam(self)

    def import_dbase_short(self, source: Literal["Stowa", "PV-tool", "Dbase"], source_dir: Path):
        """Importeert data uit de Stowa-database, de oude pv-tool of de Dbase (template) en voegt kolommen toe"""
        if source == "Dbase":
            import_dbase(self, dbase_dir=source_dir)
            return self.dbase_df
        else:
            return f"Short import only available for 'Dbase' source, not for '{source}'"

    def import_data(self, source: Literal["Stowa", "PV-tool", "Dbase"], source_dir: Path):
        """Importeert data uit de Stowa-database, de oude pv-tool of de Dbase (template) en voegt kolommen toe"""
        if source == "Stowa":
            import_stowa(self, stowa_dir=source_dir)
            self.dbase_df = self.stowa_df
        elif source == "PV-tool":
            import_pv_tool(self, pv_dir=source_dir)
            self.dbase_df = self.pv_tool
        elif source == "Dbase":
            import_dbase(self, dbase_dir=source_dir)
        self._create_dbase(source=source)
        return self.dbase_df

    def validate_data(self, export_path: Path):
        self.validation.validation_export(export_path=export_path)
        self.validation.print_critical_errors()
        return self.dbase_df

    def create_dbase_for_export(self):
        """
        Creates the Dbase-DataFrame for the export to Excel-template, maintaining the correct column order
        and preserving specified columns from the current DataFrame.
        """
        # Reorder columns for template
        base_columns = [col for col in PV_TOOL_DBASE_COLUMNS if col in self.dbase_df.columns]
        base_columns = [col for col in base_columns if col not in ANA_COLUMNS]
        ana_columns = [col for col in ANA_COLUMNS if col in self.dbase_df.columns]
        used_columns = set(base_columns + ana_columns)
        other_columns = [col for col in self.dbase_df.columns if col not in used_columns]
        final_columns = base_columns + ana_columns + other_columns

        # Save in init with reordered columns
        return self.dbase_df[final_columns].copy()

    def export_dbase_to_template(self, export_dir, export_name = "Template_PVtool5_0.xlsx"):
        """Exporteert het dbase-dataframe naar Excel-template"""
        if platform.system() == "Windows":
            return self._export_dbase_xlwings(export_dir=export_dir, export_name=export_name)
        else:
            return self._export_dbase_openpyxl(export_dir=export_dir, export_name=export_name)

    def _export_dbase_xlwings(self, export_dir, export_name="Template_PVtool5_0.xlsx"):
        """
        Exporteert het dbase-dataframe naar de Excel-template met xlwings.

        Hiermee blijven dropdowns, data-validaties en andere Excel-specifieke objecten behouden.
        """
        sheet_name = "Dbase5_0"
        start_row = 7  # Excel rij 8
        start_col = 1  # Kolom A

        export_to = str(Path(export_dir) / export_name)

        # Maak export dataframe
        export_df = self.create_dbase_for_export()

        index_col_name = export_df.index.name if export_df.index.name else "Index"

        if index_col_name in export_df.columns:
            export_df = export_df.drop(columns=[index_col_name])

        export_df.insert(0, index_col_name, export_df.index.astype(str))

        export_df = export_df.replace({pd.NA: ""})
        export_df = export_df.fillna("")

        from datetime import time

        for col in export_df.columns:
            mask = export_df[col].apply(lambda x: isinstance(x, time))
            if mask.any():
                print(f"Kolom bevat datetime.time: {col}")
                print(export_df.loc[mask, col])

        export_values = export_df.values.tolist()

        app = None
        wb = None
        opened_by_function = False

        try:
            export_path = Path(export_to).resolve()

            for app_instance in xw.apps:
                for wb_instance in app_instance.books:
                    try:
                        if Path(wb_instance.fullname).resolve() == export_path:
                            wb = wb_instance
                            print("Bestaand geopend Excel-bestand gevonden.")
                            break
                    except Exception:
                        pass

                if wb is not None:
                    break

            if wb is None:

                if export_path.exists():
                    app = xw.App(visible=False)
                    wb = app.books.open(str(export_path))
                    opened_by_function = True

                else:
                    with importlib.resources.path(
                        "pv_tool_logic.templates",
                        "Template_PVtool5_0.xlsx"
                    ) as template_path:

                        app = xw.App(visible=False)
                        wb = app.books.open(str(template_path))

                        wb.save(str(export_path))
                        opened_by_function = True

            if sheet_name not in [sheet.name for sheet in wb.sheets]:
                raise ValueError(
                    f"Sheet '{sheet_name}' bestaat niet in template!"
                )

            ws = wb.sheets[sheet_name]

            last_row = ws.used_range.last_cell.row
            last_col = ws.used_range.last_cell.column

            if last_row > start_row:
                ws.range(
                    (start_row + 1, start_col),
                    (last_row, last_col)
                ).clear_contents()

            if export_values:
                ws.range((start_row + 1, start_col)).value = export_values

            wb.save()

            print(f"DataFrame naar template geëxporteerd in {export_to}")

        finally:

            # Alleen sluiten als deze functie het bestand zelf geopend heeft
            if opened_by_function and wb is not None:
                wb.close()

            if opened_by_function and app is not None:
                app.quit()

    def _export_dbase_openpyxl(self, export_dir, export_name="Template_PVtool5_0.xlsx"):
        """Exporteert het dbase-dataframe naar de excel-template met openpyxl"""

        sheet_name = "Dbase5_0"
        start_row = 7  # Excel: rij 8
        start_col = 1  # Excel: kolom A

        export_to = os.path.join(export_dir, export_name)

        # Laad bestaand bestand om resultaten-tabbladen te behouden,
        # gebruik het lege template als het bestand nog niet bestaat
        if os.path.exists(export_to):
            wb = load_workbook(export_to)
        else:
            with importlib.resources.path("pv_tool_logic.templates", "Template_PVtool5_0.xlsx") as template_path:
                wb = load_workbook(template_path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' bestaat niet in template!")

        ws = wb[sheet_name]

        export_df = self.create_dbase_for_export()
        index_col_name = export_df.index.name if export_df.index.name else "Index"

        if index_col_name in export_df.columns:
            export_df = export_df.drop(columns=[index_col_name])

        export_df.insert(0, index_col_name, export_df.index.astype(str))

        export_df = export_df.replace({pd.NA: ""})
        export_df = export_df.fillna("")

        for i, row in enumerate(export_df.values):
            for j, value in enumerate(row):
                ws.cell(row=start_row + 1 + i, column=start_col + j, value=value)

        wb.save(export_to)
        print(f"DataFrame naar template geëxporteerd in {export_to}")