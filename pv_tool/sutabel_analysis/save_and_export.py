"""
Save and Export functies voor Sutabel-m analyse.

Deze module bevat functies voor het opslaan en exporteren van sutabel-m analyseresultaten
naar Excel en PDF formaat.
"""

from typing import TYPE_CHECKING, List
from pandas import ExcelWriter, concat, DataFrame, read_excel
from datetime import datetime
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, LongTable, PageBreak
from pv_tool.imports.excel_utils import format_excel_sheet
import numpy as np

try:
    from reportlab.platypus import Image as RLImage
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("Waarschuwing: PIL (Pillow) is niet beschikbaar. Figuren kunnen niet aan PDF worden toegevoegd.")

if TYPE_CHECKING:
    from pv_tool.sutabel_analysis.sutabel_analysis import SUTABEL


def add_sutabel_results_to_dbase(self: "SUTABEL", path: str, file_name: str = 'Template_PVtool5_0.xlsx'):
    """
    Voegt de sutabel-m analyseresultaten toe aan de database Excel-bestand.

    Parameters
    ----------
    self : sutabel
        Instantie van de sutabel analyse klasse
    path : str
        Pad naar de map waar het Excel-bestand staat
    file_name : str
        Naam van het Excel-bestand

    Returns
    -------
    DataFrame
        Bijgewerkte DataFrame met alle resultaten
    """
    file_path = f"{path}/{file_name}"

    # Ensure analysis is run
    if self.sutabel_grafiek is None or self.e_a1_sutabel is None:
        self._run_sutabel()
        if hasattr(self, 'cv_fit_kar_sutabel') and self.cv_fit_kar_sutabel is not None:
            self.get_sutabel_parameters(cv_fit_kar_sutabel=self.cv_fit_kar_sutabel)
        else:
            self.get_sutabel_parameters()

    # Expected columns structure voor sutabel resultaten
    expected_columns = [
        'PVNAAM', 'PV_REK', 'PV_TYPE_PROEF', 'PV_ANALYSE', 'PV_RESULTAAT_ID', 'PV_TYPEVERZAMELING',
        'PV_e_a1_GEM [-]', 'PV_e_a2_GEM [-]', 'PV_svgm_GEM [kPa]', 'PV_m_GEM [-]',
        'PV_a1_KAR [-]', 'PV_a2_KAR [-]', 'PV_svgm_KAR [kPa]', 'PV_m_KAR [-]',
        'PV_cv_FIT_KAR [-]', 'PV_STDEV_LOGN_cv [-]', 'PV_STEYX [-]',
        'PV_VGWNAT_GEM [kN/m3]', 'PV_VGWNAT_SD [kN/m3]', 'PV_WATERGEHALTE_GEM', 'PV_WATERGEHALTE_SD',
        'Timestamp'
    ]

    new_row = {
        'PVNAAM': self.investigation_groups[0],
        'PV_REK': self.effective_stress,
        'PV_TYPE_PROEF': self.analysis_type.split('_')[0],
        'PV_ANALYSE': '_'.join(self.analysis_type.split('_')[1:]),
        'PV_RESULTAAT_ID': f"{self.investigation_groups[0]}_{self.effective_stress}_{self.analysis_type}",
        'PV_TYPEVERZAMELING': self.alpha,
        'PV_e_a1_GEM [-]': round(self.e_a1_sutabel, 6) if self.e_a1_sutabel is not None else None,
        'PV_e_a2_GEM [-]': round(self.e_a2_sutabel, 6) if self.e_a2_sutabel is not None else None,
        'PV_svgm_GEM [kPa]': round(self.svgm_gem_sutabel, 6) if self.svgm_gem_sutabel is not None else None,
        'PV_m_GEM [-]': round(self.m_gem_sutabel, 6) if self.m_gem_sutabel is not None else None,
        'PV_a1_KAR [-]': round(self.a1_kar_sutabel, 6) if self.a1_kar_sutabel is not None else None,
        'PV_a2_KAR [-]': round(self.a2_kar_sutabel, 6) if self.a2_kar_sutabel is not None else None,
        'PV_svgm_KAR [kPa]': round(self.svgm_kar_sutabel, 6) if self.svgm_kar_sutabel is not None else None,
        'PV_m_KAR [-]': round(self.m_kar_sutabel, 6) if self.m_kar_sutabel is not None else None,
        'PV_CV_FIT_KAR [-]': round(self.cv_fit_kar_sutabel, 6) if self.cv_fit_kar_sutabel is not None else None,
        'PV_STDEV_LOGN_cv [-]': round(self.STDEV_logn_cv_sutabel, 6) if self.STDEV_logn_cv_sutabel is not None else None,
        'PV_STEYX [-]': round(self.steyx_sutabel, 6) if self.steyx_sutabel is not None else None,
        'PV_VGWNAT_GEM [kN/m3]': round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
        'PV_VGWNAT_SD [kN/m3]': round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
        'PV_WATERGEHALTE_GEM': round(self.calc_watergehalte_gem, 3) if self.calc_watergehalte_gem is not None else None,
        'PV_WATERGEHALTE_SD': round(self.calc_watergehalte_sd, 3) if self.calc_watergehalte_sd is not None else None,
        'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    workbook = load_workbook(file_path)

    if 'Resultaten SU-tabel - m' in workbook.sheetnames:
        print('Tabblad Resultaten SU-tabel - m in dbase excel bestaat al en wordt aangevuld')
        df_existing = read_excel(file_path, sheet_name='Resultaten SU-tabel - m')
        # Filter out empty rows and ensure consistent types before concatenation
        df_existing = df_existing.dropna(how='all')
        # Ensure column headers are strings
        df_existing.columns = df_existing.columns.astype(str)
        new_row_df = DataFrame([new_row], columns=df_existing.columns)
        df_updated = concat([df_existing, new_row_df], ignore_index=True)
    else:
        print('Tabblad Resultaten SU-tabel - m in dbase excel bestaat nog niet en wordt aangemaakt')
        df_updated = DataFrame([new_row], columns=expected_columns)

    # Ensure all column headers are strings
    df_updated.columns = df_updated.columns.astype(str)

    # Write data to Excel
    with ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_updated.to_excel(writer, sheet_name='Resultaten SU-tabel - m', index=False)

    # Format the Excel sheet
    format_excel_sheet(
        file_path=file_path,
        sheet_name='Resultaten SU-tabel - m',
        num_columns=df_updated.shape[1],
        num_rows=df_updated.shape[0],
        table_name='ResultatenSUTabelMTable',
        index=False
    )

    print(f"Sutabel resultaten toegevoegd aan database: {file_path}")
    return df_updated


def _create_sutabel_input_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de invoerselectie informatie voor sutabel analyse.

    Returns
    -------
    Table
        ReportLab tabel object met de invoerselectie informatie
    """
    if self.shansep_data_df_sutabel is None:
        return Table([['Geen invoerdata beschikbaar']], hAlign='LEFT')

    # Selecteer relevante kolommen
    columns_base = ['PV_NAAM', 'BORING_POSITIE', 'MONSTER_NIVEAU_NAP_VANAF', 'MONSTER_NIVEAU_NAP_TOT',
                    'ANA_TERREINSPANNING', 'ANA_GRENSSPANNING_REKEN', 'ANA_POP_VELD']

    if self.analysis_type.startswith('TXT'):
        columns_extra = ['TXT_SS_VOLUMEGEWICHT_NAT', 'TXT_SS_VOLUMEGEWICHT_DRG', 'TXT_SS_WATERGEHALTE_VOOR']
    else:
        columns_extra = ['DSS_VOLUMEGEWICHT_NAT', 'DSS_VOLUMEGEWICHT_DRG', 'DSS_WATERGEHALTE_VOOR']

    # Check welke kolommen bestaan in de dataframe
    available_columns = [col for col in columns_base + columns_extra if col in self.total_shansep_data_df.columns]

    if not available_columns:
        return Table([['Geen invoerdata kolommen beschikbaar']], hAlign='LEFT')

    table_df = self.total_shansep_data_df[available_columns].copy()

    # Add additional columns from shansep_data_df_sutabel
    additional_columns = ['S\'v', 'Su', 'consolidatietype']
    for col in additional_columns:
        if col in self.shansep_data_df_sutabel.columns and col not in table_df.columns:
            # Zorg ervoor dat de indices overeenkomen
            if len(table_df) == len(self.shansep_data_df_sutabel):
                table_df[col] = self.shansep_data_df_sutabel[col].values

    # Round numeric columns
    for col in table_df.columns:
        if table_df[col].dtype in ['float64', 'float32']:
            table_df[col] = table_df[col].round(3)

    # Maak header met line breaks voor lange kolomnamen
    from reportlab.platypus import Paragraph
    header_paragraphs = []
    for col in table_df.columns:
        # Splits lange kolomnamen op underscores voor betere leesbaarheid
        col_display = col.replace('_', '_<br/>')
        header_paragraphs.append(Paragraph(f'<font size=7>{col_display}</font>', getSampleStyleSheet()['Normal']))

    # Converteer DataFrame naar lijst
    data = table_df.values.tolist()
    # Voeg header toe
    t_data = [header_paragraphs] + data

    # Bereken kolom breedtes - geef meer ruimte aan belangrijke kolommen
    num_cols = len(table_df.columns)
    col_widths = [60 if i < 2 else 50 for i in range(num_cols)]  # Eerste 2 kolommen iets breder

    t1 = LongTable(t_data, repeatRows=1, hAlign='LEFT', colWidths=col_widths)
    t1.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 0), (-1, 0), [colors.lightgrey]),
    ]))
    return t1


def _create_sutabel_parameters_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de sutabel-m parameters.

    Returns
    -------
    Table
        ReportLab tabel object met de sutabel-m parameters
    """
    parameters: List[list] = []

    # Voeg sutabel parameters toe
    if hasattr(self, 'e_a1_sutabel') and self.e_a1_sutabel is not None:
        parameters.append(['e_a1 (snijpunt gemiddeld) [-]', f"{self.e_a1_sutabel:.6f}"])
    if hasattr(self, 'e_a2_sutabel') and self.e_a2_sutabel is not None:
        parameters.append(['e_a2 (helling gemiddeld) [-]', f"{self.e_a2_sutabel:.6f}"])
    if hasattr(self, 'svgm_gem_sutabel') and self.svgm_gem_sutabel is not None:
        parameters.append(['svgm_gem [kPa]', f"{self.svgm_gem_sutabel:.6f}"])
    if hasattr(self, 'm_gem_sutabel') and self.m_gem_sutabel is not None:
        parameters.append(['m_gem [-]', f"{self.m_gem_sutabel:.6f}"])

    if hasattr(self, 'a1_kar_sutabel') and self.a1_kar_sutabel is not None:
        parameters.append(['a1_kar (snijpunt karakteristiek) [-]', f"{self.a1_kar_sutabel:.6f}"])
    if hasattr(self, 'a2_kar_sutabel') and self.a2_kar_sutabel is not None:
        parameters.append(['a2_kar (helling karakteristiek) [-]', f"{self.a2_kar_sutabel:.6f}"])
    if hasattr(self, 'svgm_kar_sutabel') and self.svgm_kar_sutabel is not None:
        parameters.append(['svgm_kar [kPa]', f"{self.svgm_kar_sutabel:.6f}"])
    if hasattr(self, 'm_kar_sutabel') and self.m_kar_sutabel is not None:
        parameters.append(['m_kar [-]', f"{self.m_kar_sutabel:.6f}"])

    if hasattr(self, 'cv_fit_kar_sutabel') and self.cv_fit_kar_sutabel is not None:
        parameters.append(['cv_fit_kar [-]', f"{self.cv_fit_kar_sutabel:.6f}"])
    if hasattr(self, 'STDEV_logn_cv_sutabel') and self.STDEV_logn_cv_sutabel is not None:
        parameters.append(['STDEV lognormaal [-]', f"{self.STDEV_logn_cv_sutabel:.6f}"])
    if hasattr(self, 'steyx_sutabel') and self.steyx_sutabel is not None:
        parameters.append(['STEYX [-]', f"{self.steyx_sutabel:.6f}"])

    parameters.append(['Type verzameling: lokaal = 1.0; regionaal = 0.75', f"{self.alpha:.2f}"])

    if hasattr(self, 'calc_vgwnat_gem') and self.calc_vgwnat_gem is not None:
        parameters.append(['VGW nat gemiddeld [kN/m3]', f"{self.calc_vgwnat_gem:.3f}"])
    if hasattr(self, 'calc_watergehalte_gem') and self.calc_watergehalte_gem is not None:
        parameters.append(['Watergehalte gemiddeld', f"{self.calc_watergehalte_gem:.3f}"])

    t = Table([['Parameter', 'Waarde']] + parameters, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    return t


def _create_sutabel_grafiek_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de sutabel grafiek data (su_gem en su_kar lijnen).

    Returns
    -------
    Table
        ReportLab tabel object met de sutabel grafiek data
    """
    if self.sutabel_grafiek is None:
        return Table([['Geen sutabel grafiek data beschikbaar']], hAlign='LEFT')

    # Maak een kopie en rond af
    df = self.sutabel_grafiek.copy()
    for col in df.columns:
        if df[col].dtype in ['float64', 'float32']:
            df[col] = df[col].round(3)

    # Converteer naar tabel formaat
    header = df.columns.tolist()
    data = df.values.tolist()
    t_data = [header] + data

    t = Table(t_data, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    return t


def _create_su_fit_cv_table(self: "SUTABEL") -> Table:
    """
    Maakt een tabel met de su fit constante cv data.

    Returns
    -------
    Table
        ReportLab tabel object met de su fit cv data
    """
    if self.su_fit_constante_cv is None:
        return Table([['Geen cv fit data beschikbaar (cv_fit_kar niet opgegeven)']], hAlign='LEFT')

    # Maak een kopie en rond af
    df = self.su_fit_constante_cv.copy()
    for col in df.columns:
        if df[col].dtype in ['float64', 'float32']:
            df[col] = df[col].round(3)

    # Converteer naar tabel formaat
    header = df.columns.tolist()
    data = df.values.tolist()
    t_data = [header] + data

    t = Table(t_data, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    return t


def save_sutabel_to_pdf(self: "SUTABEL", path: str, cv_fit_kar: float = None) -> str:
    """
    Slaat de sutabel-m analyseresultaten op in een PDF-document.

    De PDF bevat:
    - Titel met analysedetails
    - Beide overzichtsfiguren (ln(s'v)-ln(su) en s'v-su)
    - Tabel met parameters
    - Tabel met sutabel grafiek data (su_gem en su_kar)
    - Tabel met su fit constante cv data (indien van toepassing)
    - Tabel met invoerselectie informatie

    Parameters
    ----------
    path : str
        Map locatie waar het PDF-bestand moet worden opgeslagen
    cv_fit_kar : float, optioneel
        Coefficient of Variation voor de fit

    Returns
    -------
    str
        Het absolute bestandspad van het aangemaakte PDF-bestand
    """
    # Maak titel en bestandsnaam
    title = f"Sutabel-m analyse met {self.effective_stress} op {self.investigation_groups[0]}"
    file_name = f"sutabel_pdf_export_{self.investigation_groups[0]}_{self.analysis_type}_{str(self.effective_stress).replace('%', 'procent_').replace(' ', '')}.pdf"
    file_path = f"{path}/{file_name}"

    # Ensure analysis is run with cv parameter if provided
    self._run_sutabel()
    if cv_fit_kar is not None:
        self.get_sutabel_parameters(cv_fit_kar_sutabel=cv_fit_kar)
    elif hasattr(self, 'cv_fit_kar_sutabel') and self.cv_fit_kar_sutabel is not None:
        self.get_sutabel_parameters(cv_fit_kar_sutabel=self.cv_fit_kar_sutabel)
    else:
        self.get_sutabel_parameters()

    # Bereken sutabel grafiek data
    if self.sutabel_grafiek is None:
        self.calculate_sutabel_grafiek()

    # Maak het PDF document
    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    # Voeg alleen styles toe die nog niet bestaan
    if 'Left' not in styles:
        styles.add(ParagraphStyle(name='Left', parent=styles['Normal'], alignment=TA_LEFT))
    if 'TitleLeft' not in styles:
        styles.add(ParagraphStyle(name='TitleLeft', parent=styles['Title'], alignment=TA_LEFT))
    if 'Heading3' not in styles:
        styles.add(ParagraphStyle(name='Heading3', parent=styles['Heading2'], alignment=TA_LEFT))

    story = []

    # Voeg titel toe
    story.append(Paragraph(title, styles['TitleLeft']))
    story.append(Spacer(width=1, height=12))

    if not PIL_AVAILABLE:
        story.append(Paragraph("Figuren kunnen niet worden toegevoegd: PIL (Pillow) niet beschikbaar", styles['Normal']))
        story.append(Spacer(width=1, height=12))
    else:
        # Eerste figuur: ln(s'v) vs ln(su) plot
        try:
            fig_path1 = f"{path}/temp_sutabel_plot1.png"
            self.show_title = False

            # Genereer figuur
            self.figure = None
            self.set_figure_ln_sv_ln_su_sutabel()

            if hasattr(self, 'figure') and self.figure is not None:
                fig_width = 1280
                fig_height = 720
                self.figure.write_image(fig_path1, width=fig_width, height=fig_height, scale=4, format="png")

                # Bereken figuurgrootte voor pagina
                title_and_heading_height = 100
                available_height = doc.height - title_and_heading_height

                # Laad PNG en bepaal pixelafmetingen
                with PILImage.open(fig_path1) as im:
                    img_width_px, img_height_px = im.size

                # Bereken optimale grootte die op de pagina past
                max_width_pt = doc.width * 0.95
                aspect = img_height_px / img_width_px

                # Probeer eerst op basis van breedte
                img_width_pt = min(max_width_pt, doc.width)
                img_height_pt = img_width_pt * aspect

                # Controleer of het past in de beschikbare hoogte
                if img_height_pt > available_height:
                    img_height_pt = available_height * 0.95
                    img_width_pt = img_height_pt / aspect

                # Maak ReportLab Image aan
                img1 = RLImage(fig_path1)
                img1.drawWidth = img_width_pt
                img1.drawHeight = img_height_pt
                img1.hAlign = 'LEFT'

                story.append(Paragraph("Figuur 1: ln(s'v) vs ln(su)", styles['Heading2']))
                story.append(Spacer(width=1, height=6))
                story.append(img1)
                story.append(PageBreak())
            else:
                story.append(Paragraph("Figuur 1 kon niet worden gegenereerd", styles['Normal']))
                story.append(Spacer(width=1, height=12))
        except Exception as e:
            story.append(Paragraph(f"Fout bij genereren Figuur 1: {str(e)}", styles['Normal']))
            story.append(Spacer(width=1, height=12))

        # Tweede figuur: s'v vs su plot
        try:
            fig_path2 = f"{path}/temp_sutabel_plot2.png"
            self.show_title = False

            # Genereer figuur
            self.figure = None
            self.set_figure_sv_su_sutabel()

            if hasattr(self, 'figure') and self.figure is not None:
                fig_width = 1280
                fig_height = 720
                self.figure.write_image(fig_path2, width=fig_width, height=fig_height, scale=4, format="png")

                # Bereken figuurgrootte
                with PILImage.open(fig_path2) as im:
                    img_width_px, img_height_px = im.size

                max_width_pt = doc.width * 0.95
                aspect = img_height_px / img_width_px

                img_width_pt = min(max_width_pt, doc.width)
                img_height_pt = img_width_pt * aspect

                available_height = doc.height - 50  # Ruimte voor heading

                if img_height_pt > available_height:
                    img_height_pt = available_height * 0.95
                    img_width_pt = img_height_pt / aspect

                # Maak ReportLab Image aan
                img2 = RLImage(fig_path2)
                img2.drawWidth = img_width_pt
                img2.drawHeight = img_height_pt
                img2.hAlign = 'LEFT'

                story.append(Paragraph("Figuur 2: s'v vs su", styles['Heading2']))
                story.append(Spacer(width=1, height=6))
                story.append(img2)
                story.append(PageBreak())
            else:
                story.append(Paragraph("Figuur 2 kon niet worden gegenereerd", styles['Normal']))
                story.append(Spacer(width=1, height=12))
        except Exception as e:
            story.append(Paragraph(f"Fout bij genereren Figuur 2: {str(e)}", styles['Normal']))
            story.append(Spacer(width=1, height=12))

    # Voeg parameters tabel toe
    story.append(Paragraph("Parameters", styles['Heading2']))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_parameters_table(self))
    story.append(Spacer(width=1, height=12))

    # Voeg sutabel grafiek data toe
    story.append(Paragraph("Sutabel grafiek data (su_gem en su_kar lijnen)", styles['Heading2']))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_grafiek_table(self))
    story.append(Spacer(width=1, height=12))

    # Voeg su fit cv data toe (indien beschikbaar)
    if self.su_fit_constante_cv is not None:
        story.append(Paragraph("Su fit met constante cv data", styles['Heading2']))
        story.append(Spacer(width=1, height=6))
        story.append(_create_su_fit_cv_table(self))
        story.append(Spacer(width=1, height=12))

    # Voeg input tabel toe op nieuwe pagina
    story.append(PageBreak())
    story.append(Paragraph("Invoerselectie", styles['Heading2']))
    story.append(Spacer(width=1, height=6))
    story.append(_create_sutabel_input_table(self))

    # Bouw PDF
    try:
        doc.build(story)
        print(f"Sutabel PDF export voltooid: {file_path}")
        return file_path
    except Exception as e:
        print(f"Fout bij maken PDF: {e}")
        raise

